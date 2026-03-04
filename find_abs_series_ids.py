#!/usr/bin/env python3
"""
find_abs_series_ids.py
======================
Diagnostic tool — prints all Series IDs in an ABS XLSX file
alongside their description and series type.

Run this to find the correct IDs for HSR and TOT after an ID change.

Usage:
    python find_abs_series_ids.py /path/to/5206001_Key_Aggregates.xlsx

    # Or with search filter:
    python find_abs_series_ids.py /path/to/5206001_Key_Aggregates.xlsx --search "saving"
    python find_abs_series_ids.py /path/to/5206001_Key_Aggregates.xlsx --search "trade"
"""

import argparse
import sys
from pathlib import Path

import openpyxl


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="Path to ABS XLSX file")
    parser.add_argument("--search", default=None,
                        help="Filter: only show rows containing this string (case-insensitive)")
    args = parser.parse_args()

    path = Path(args.path)
    if not path.exists():
        print(f"ERROR: File not found: {path}")
        sys.exit(1)

    print(f"Loading {path.name} ...")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet = "Data1" if "Data1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))

    # ABS standard metadata rows
    META = {
        0: "Description",
        1: "Unit",
        2: "Series Type",
        3: "Data Type",
        4: "Frequency",
        9: "Series ID",
    }

    sid_row = next((i for i, r in enumerate(rows) if r and r[0] == "Series ID"), None)
    if sid_row is None:
        print("ERROR: No 'Series ID' row found")
        sys.exit(1)

    series_ids   = [str(v).strip() if v else "" for v in rows[sid_row]]
    descriptions = [str(v).strip() if v else "" for v in rows[0]]
    series_types = [str(v).strip() if v else "" for v in rows[2]]
    units        = [str(v).strip() if v else "" for v in rows[1]]

    # Count data rows
    data_rows = rows[sid_row + 1:]
    n_data = sum(1 for r in data_rows if r and r[0] is not None)

    print(f"\nSheet: {sheet}")
    print(f"Series ID row: {sid_row}")
    print(f"Total series: {len([s for s in series_ids if s and s != 'Series ID'])}")
    print(f"Data rows: {n_data}")
    print()

    # Print all series
    search = args.search.lower() if args.search else None
    print(f"{'Series ID':<16} {'Type':<22} {'Unit':<18} {'Description'}")
    print("-" * 120)

    found = []
    for i, sid in enumerate(series_ids):
        if not sid or sid == "Series ID":
            continue
        desc = descriptions[i] if i < len(descriptions) else ""
        stype = series_types[i] if i < len(series_types) else ""
        unit = units[i] if i < len(units) else ""

        if search and search not in desc.lower() and search not in sid.lower():
            continue

        print(f"{sid:<16} {stype:<22} {unit:<18} {desc[:70]}")
        found.append((sid, stype, unit, desc))

    if search:
        print(f"\nFound {len(found)} matches for '{args.search}'")
    else:
        print(f"\nTotal series listed: {len(found)}")

    # Highlight likely HSR and TOT candidates
    print("\n" + "=" * 60)
    print("LIKELY MATCHES FOR HSR AND TOT:")
    print("=" * 60)
    keywords = {
        "HSR": ["saving ratio", "household saving", "saving to income"],
        "TOT": ["terms of trade", "terms-of-trade"],
    }
    for target, kws in keywords.items():
        print(f"\n{target}:")
        for i, sid in enumerate(series_ids):
            if not sid or sid == "Series ID":
                continue
            desc = descriptions[i].lower() if i < len(descriptions) else ""
            stype = series_types[i] if i < len(series_types) else ""
            if any(kw in desc for kw in kws) and "Seasonally Adjusted" in stype:
                print(f"  ✓ {sid:<16} {stype:<22} {descriptions[i][:70]}")


if __name__ == "__main__":
    main()
