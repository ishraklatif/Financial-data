#!/usr/bin/env python3
"""
fetch_abs.py
============
Parses ABS National Accounts data from XLSX table downloads.

The ABS previously provided bulk SDMX downloads (all.xml / ANA_AGG.zip)
but removed this from release pages. This script reads the XLSX tables
instead, which are always available on every release page.

Files required (place in data/raw/abs/downloads/):
  5206001_Key_Aggregates.xlsx   — Table 1: Key National Accounts Aggregates
                                  Contains: GDP, GDP_PCA, HSR, TOT
  634501.xlsx                   — WPI: Wage Price Index
                                  Contains: WPI

How to download Table 1:
  1. Go to the latest ANA release page:
     https://www.abs.gov.au/statistics/economy/national-accounts/
     australian-national-accounts-national-income-expenditure-and-product/latest-release
  2. Scroll down to the data tables section
  3. Click "Table 1. Key National Accounts Aggregates" → Download XLSX
  4. Save as data/raw/abs/downloads/5206001_Key_Aggregates.xlsx

How to download WPI (634501.xlsx):
  Already in your downloads folder. Re-download from:
  https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/
  wage-price-index-australia/latest-release
  Download "Table 1. Total Hourly Rates of Pay Excluding Bonuses"

ABS XLSX format (consistent across all publications):
  Row 0:  Series description
  Row 1:  Unit
  Row 2:  Series Type (Seasonally Adjusted / Original / Trend)
  Row 3:  Data Type
  Row 4:  Frequency
  Row 5:  Collection Month
  Row 6:  Series Start
  Row 7:  Series End
  Row 8:  No. Obs
  Row 9:  Series ID  ← column headers
  Row 10+: data (col A = datetime, col B+ = values)

Target series:
  GDP       A2304402X  Gross domestic product, Chain volume, SA ($m)
  GDP_PCA   A2304404C  GDP per capita, Chain volume, SA
  HSR       A2304418V  Household saving ratio, SA (%)
  TOT       A2304451K  Terms of trade index, SA
  WPI       A2713849C  WPI total hourly rates, all sectors, SA (index)

NOTE on Series IDs:
  ABS Series IDs are stable across releases but the ABS occasionally
  adds new series or revises IDs after major methodological changes.
  If a series is not found, the script prints available IDs and exits
  with an error so you can update the TARGET_SERIES dict below.

Usage:
    python -m scripts.fetch.abs.fetch_abs
    python -m scripts.fetch.abs.fetch_abs --t1 /path/to/table1.xlsx
    python -m scripts.fetch.abs.fetch_abs --verify

Output:
    data/raw/abs/gdp/GDP.parquet
    data/raw/abs/gdp/GDP_PCA.parquet
    data/raw/abs/gdp/HSR.parquet
    data/raw/abs/gdp/TOT.parquet
    data/raw/abs/wpi/WPI.parquet
    data/raw/abs/_fetch_summary.json
"""

import argparse
import json
import logging
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import yaml
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv(PROJECT_ROOT / ".env")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(PROJECT_ROOT / "logs" / "fetch_abs.log"),
    ],
)
log = logging.getLogger(__name__)

CONFIG_PATH   = PROJECT_ROOT / "config" / "data.yaml"
DOWNLOADS_DIR = PROJECT_ROOT / "data" / "raw" / "abs" / "downloads"

# ─────────────────────────────────────────────────────────────────────────────
# Target series configuration
# ─────────────────────────────────────────────────────────────────────────────

# Table 1 — Key National Accounts Aggregates (5206001_Key_Aggregates.xlsx)
# Series type filter: "Seasonally Adjusted" only
TABLE1_SERIES = [
    {
        "canonical_name": "GDP",
        "series_id":      "A2304402X",
        "description":    "GDP chain volume, seasonally adjusted (AUD millions)",
        "out_dir":        "gdp",
        "frequency":      "quarterly",
    },
    {
        "canonical_name": "GDP_PCA",
        "series_id":      "A2304404C",
        "description":    "GDP per capita, seasonally adjusted",
        "out_dir":        "gdp",
        "frequency":      "quarterly",
    },
    {
        "canonical_name": "HSR",
        "series_id":      "A2323382F",
        "description":    "Household saving ratio, seasonally adjusted (%) — stored as proportion in ABS, multiplied by 100",
        "out_dir":        "gdp",
        "frequency":      "quarterly",
        "transform":      "x100",   # ABS stores as proportion (0.064), convert to percent (6.4)
    },
    {
        "canonical_name": "TOT",
        "series_id":      "A2304200A",
        "description":    "Terms of trade index, seasonally adjusted (index numbers)",
        "out_dir":        "gdp",
        "frequency":      "quarterly",
    },
]

# WPI — Wage Price Index (634501.xlsx)
WPI_SERIES = {
    "canonical_name": "WPI",
    "series_id":      "A2713849C",
    "description":    "Wage price index — total hourly rates all sectors SA (quarterly)",
    "out_dir":        "wpi",
    "frequency":      "quarterly",
}

# Fallback series IDs if primary not found (ABS revises IDs occasionally)
# Format: {primary_id: [fallback1, fallback2, ...]}
FALLBACK_IDS = {
    "A2304402X": ["A2304370T", "A2304371V"],  # GDP CVM SA
    "A2304404C": ["A2304372W", "A2304404C"],  # GDP per capita SA
    "A2323382F": ["A2304418V", "A2304425C"],  # HSR SA (old ID: A2304418V)
    "A2304200A": ["A2304451K", "A2304449A"],  # Terms of trade index SA (old ID: A2304451K)
}


def load_config() -> dict:
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


# ─────────────────────────────────────────────────────────────────────────────
# XLSX parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_abs_xlsx(path: Path, series_id: str,
                   canonical_name: str) -> pd.DataFrame:
    """
    Parse an ABS XLSX file (Data1 sheet) for a specific Series ID.

    ABS XLSX format:
      Row 9  (index 9):  Series ID header row
      Row 10+ (index 10+): date in col A, values in subsequent cols
      Col A dates are datetime objects or date strings

    Returns DataFrame with columns: date, value, series_id, series_name
    Raises ValueError if series_id not found.
    """
    log.info(f"  Parsing {path.name} for {canonical_name} ({series_id}) ...")

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open {path}: {e}")

    # Try Data1 sheet first, then first available sheet
    sheet_name = "Data1" if "Data1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find Series ID row (row 9 in ABS standard, but search to be safe)
    sid_row_idx = next(
        (i for i, row in enumerate(rows) if row and row[0] == "Series ID"),
        None
    )
    if sid_row_idx is None:
        raise ValueError(f"No 'Series ID' row found in {path.name}")

    series_ids = [str(v).strip() if v is not None else "" for v in rows[sid_row_idx]]

    # Try primary ID, then fallbacks
    ids_to_try = [series_id] + FALLBACK_IDS.get(series_id, [])
    col_idx = None
    used_id = None
    for sid in ids_to_try:
        if sid in series_ids:
            col_idx = series_ids.index(sid)
            used_id = sid
            break

    if col_idx is None:
        available = [s for s in series_ids if s and s != "Series ID"]
        raise ValueError(
            f"Series ID '{series_id}' not found in {path.name}.\n"
            f"  Tried: {ids_to_try}\n"
            f"  Available IDs: {available[:20]}\n"
            f"  → Update TARGET_SERIES in fetch_abs.py with the correct ID."
        )

    if used_id != series_id:
        log.warning(f"  Used fallback ID '{used_id}' instead of '{series_id}'")

    # Parse data rows
    records = []
    for row in rows[sid_row_idx + 1:]:
        if not row or row[0] is None:
            continue

        # Date: ABS provides datetime objects from openpyxl
        date_val = row[0]
        if isinstance(date_val, datetime):
            dt = pd.Timestamp(date_val)
        elif isinstance(date_val, date):
            dt = pd.Timestamp(date_val)
        elif isinstance(date_val, str):
            dt = None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%b-%Y"):
                try:
                    dt = pd.Timestamp(datetime.strptime(date_val.strip(), fmt))
                    break
                except ValueError:
                    continue
            if dt is None:
                continue
        else:
            continue

        # Value
        val = float("nan")
        if col_idx < len(row) and row[col_idx] is not None:
            try:
                val = float(row[col_idx])
            except (ValueError, TypeError):
                pass

        records.append({"date": dt, "value": val})

    if not records:
        raise ValueError(f"No data rows parsed for {series_id} in {path.name}")

    df = (
        pd.DataFrame(records)
        .assign(
            series_id=used_id,
            series_name=canonical_name,
            frequency="quarterly",
            source="ABS",
        )
        .dropna(subset=["date"])
        .sort_values("date")
        .drop_duplicates("date")
        .reset_index(drop=True)
    )

    log.info(
        f"  → {len(df)} rows | "
        f"{df['date'].min().date()} → {df['date'].max().date()} | "
        f"missing={df['value'].isna().sum()}"
    )
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Save + manifest
# ─────────────────────────────────────────────────────────────────────────────

def save(canonical_name: str, description: str, frequency: str,
         df: pd.DataFrame, start: str, out_dir: Path) -> dict:
    df = df[df["date"] >= start].copy().reset_index(drop=True)

    issues    = []
    n_missing = int(df["value"].isna().sum())
    if n_missing > 0:
        issues.append({"code": "MISSING_VALUES", "count": n_missing})

    # Staleness check for quarterly series
    if len(df) > 0:
        last_date  = df["date"].max().date()
        stale_days = (date.today() - last_date).days
        threshold  = 120  # quarterly: allow up to 4 months lag
        if stale_days > threshold:
            issues.append({
                "code":   "STALE_DATA",
                "detail": f"Last observation {last_date} is {stale_days} days ago "
                          f"(threshold: {threshold})",
            })

    status = "WARN" if issues else "OK"

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{canonical_name}.parquet"
    df.to_parquet(out_path, index=False)

    result = {
        "canonical_name": canonical_name,
        "description":    description,
        "status":         status,
        "issues":         issues,
        "rows":           len(df),
        "date_min":       str(df["date"].min().date()) if len(df) else None,
        "date_max":       str(df["date"].max().date()) if len(df) else None,
        "missing_pct":    round(n_missing / len(df) * 100, 2) if len(df) else 0,
        "frequency":      frequency,
        "output":         str(out_path),
        "columns":        list(df.columns),
    }

    manifest_path = out_dir / f"{canonical_name}_manifest.json"
    with open(manifest_path, "w") as f:
        json.dump({**result, "fetched_at": datetime.utcnow().isoformat()}, f, indent=2)

    icon = "✓" if status == "OK" else "⚠"
    log.info(
        f"  {icon} [{canonical_name}] {status} | rows={len(df)} | "
        f"missing={n_missing} | "
        f"{result['date_min']} → {result['date_max']}"
    )
    return result


def error_result(canonical_name: str, detail: str) -> dict:
    log.error(f"  ✗ [{canonical_name}] ERROR: {detail}")
    return {
        "canonical_name": canonical_name,
        "status":         "ERROR",
        "issues":         [{"code": "ERROR", "detail": str(detail)[:400]}],
        "rows":           0,
        "date_min":       None,
        "date_max":       None,
    }


def verify(abs_dir: Path) -> None:
    """Quick verification of all ABS parquet outputs."""
    targets = [
        ("gdp",  "GDP"),
        ("gdp",  "GDP_PCA"),
        ("gdp",  "HSR"),
        ("gdp",  "TOT"),
        ("wpi",  "WPI"),
    ]
    log.info("=" * 55)
    log.info("VERIFICATION — ABS parquet outputs")
    all_ok = True
    for subdir, name in targets:
        path = abs_dir / subdir / f"{name}.parquet"
        if not path.exists():
            log.warning(f"  ✗ {name}: NOT FOUND")
            all_ok = False
            continue
        df = pd.read_parquet(path)
        df["date"] = pd.to_datetime(df["date"])
        stale = (date.today() - df["date"].max().date()).days
        log.info(
            f"  {'✓' if stale <= 120 else '⚠'} {name:<12} "
            f"rows={len(df):3d} | "
            f"{str(df['date'].min().date()):12} → {str(df['date'].max().date()):12} | "
            f"stale={stale}d"
        )
    log.info("=" * 55)


# ─────────────────────────────────────────────────────────────────────────────
# File resolution
# ─────────────────────────────────────────────────────────────────────────────

def find_table1(arg_path: str | None) -> Path | None:
    """
    Find Table 1 XLSX in priority order:
      1. CLI --t1 argument
      2. data/raw/abs/downloads/5206001_Key_Aggregates.xlsx
      3. Any file in downloads/ matching '5206001*' or 'table1*' or '*Key_Aggregate*'
    """
    if arg_path:
        p = Path(arg_path)
        if p.exists():
            return p
        log.warning(f"CLI path not found: {p}")

    # Exact name
    exact = DOWNLOADS_DIR / "5206001_Key_Aggregates.xlsx"
    if exact.exists():
        return exact

    # Fuzzy match
    patterns = ["5206001*.xlsx", "table1*.xlsx", "*Key_Aggregate*.xlsx",
                "*key_aggregate*.xlsx", "*Table_1*.xlsx"]
    for pattern in patterns:
        matches = list(DOWNLOADS_DIR.glob(pattern))
        if matches:
            log.info(f"Found Table 1 via pattern '{pattern}': {matches[0].name}")
            return matches[0]

    return None


def find_wpi(arg_path: str | None) -> Path | None:
    """Find WPI XLSX."""
    if arg_path:
        p = Path(arg_path)
        if p.exists():
            return p

    exact = DOWNLOADS_DIR / "634501.xlsx"
    if exact.exists():
        return exact

    patterns = ["634501*.xlsx", "*wage_price*.xlsx", "*WPI*.xlsx", "*wpi*.xlsx"]
    for pattern in patterns:
        matches = list(DOWNLOADS_DIR.glob(pattern))
        if matches:
            return matches[0]

    return None


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--t1",  default=None,
                        help="Path to Table 1 XLSX (5206001_Key_Aggregates.xlsx)")
    parser.add_argument("--wpi", default=None,
                        help="Path to WPI XLSX (634501.xlsx)")
    parser.add_argument("--verify", action="store_true",
                        help="Verify existing parquet outputs without re-parsing")
    args = parser.parse_args()

    cfg   = load_config()
    start = cfg["data"]["start_date"]

    abs_dir = PROJECT_ROOT / "data" / "raw" / "abs"

    if args.verify:
        verify(abs_dir)
        return

    log.info("=" * 60)
    log.info("fetch_abs.py — ABS National Accounts (XLSX parser)")
    log.info(f"Start date    : {start}")
    log.info(f"Downloads dir : {DOWNLOADS_DIR}")
    log.info("=" * 60)

    results = []

    # ── Table 1: GDP / GDP_PCA / HSR / TOT ───────────────────────────────────
    t1_path = find_table1(args.t1)
    if t1_path:
        log.info(f"Table 1: {t1_path.name}")
        for spec in TABLE1_SERIES:
            try:
                df = parse_abs_xlsx(t1_path, spec["series_id"],
                                    spec["canonical_name"])
                # Apply transform if specified
                if spec.get("transform") == "x100":
                    df["value"] = df["value"] * 100
                    log.info(f"  Applied x100 transform to {spec['canonical_name']} (proportion → percent)")
                out_dir = abs_dir / spec["out_dir"]
                results.append(
                    save(spec["canonical_name"], spec["description"],
                         spec["frequency"], df, start, out_dir)
                )
            except Exception as e:
                results.append(error_result(spec["canonical_name"], str(e)))
    else:
        log.warning(
            "Table 1 XLSX not found. "
            "Download from the ANA release page and save as:\n"
            f"  {DOWNLOADS_DIR / '5206001_Key_Aggregates.xlsx'}\n"
            "URL: https://www.abs.gov.au/statistics/economy/national-accounts/"
            "australian-national-accounts-national-income-expenditure-and-product/"
            "latest-release\n"
            "Click: 'Table 1. Key National Accounts Aggregates' → Download XLSX"
        )
        for spec in TABLE1_SERIES:
            results.append(error_result(
                spec["canonical_name"],
                "Table 1 XLSX not found — see log for download instructions"
            ))

    # ── WPI ──────────────────────────────────────────────────────────────────
    wpi_path = find_wpi(args.wpi)
    if wpi_path:
        log.info(f"WPI: {wpi_path.name}")
        try:
            df = parse_abs_xlsx(wpi_path, WPI_SERIES["series_id"],
                                WPI_SERIES["canonical_name"])
            out_dir = abs_dir / WPI_SERIES["out_dir"]
            results.append(
                save(WPI_SERIES["canonical_name"], WPI_SERIES["description"],
                     WPI_SERIES["frequency"], df, start, out_dir)
            )
        except Exception as e:
            results.append(error_result(WPI_SERIES["canonical_name"], str(e)))
    else:
        log.warning(
            "WPI XLSX (634501.xlsx) not found. "
            f"Place in: {DOWNLOADS_DIR}"
        )
        results.append(error_result(
            "WPI", "WPI XLSX not found — place 634501.xlsx in abs/downloads/"
        ))

    # ── Summary ───────────────────────────────────────────────────────────────
    n_ok   = sum(1 for r in results if r["status"] == "OK")
    n_warn = sum(1 for r in results if r["status"] == "WARN")
    n_err  = sum(1 for r in results if r["status"] == "ERROR")

    summary = {
        "generated_at": datetime.utcnow().isoformat(),
        "start":        start,
        "total":        len(results),
        "ok":           n_ok,
        "warn":         n_warn,
        "error":        n_err,
        "series":       results,
    }

    summary_path = abs_dir / "_fetch_summary.json"
    abs_dir.mkdir(parents=True, exist_ok=True)
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)

    log.info("=" * 60)
    log.info("FETCH COMPLETE")
    log.info(f"  OK    : {n_ok}")
    log.info(f"  WARN  : {n_warn}")
    log.info(f"  ERROR : {n_err}")
    log.info(f"  Summary → {summary_path}")
    log.info("=" * 60)

    if n_err > 0:
        errored = [r["canonical_name"] for r in results if r["status"] == "ERROR"]
        log.error(f"ERRORS in: {errored}")
        sys.exit(1)


if __name__ == "__main__":
    main()